import io
from collections import defaultdict, OrderedDict

import numpy as np
import pandas as pd
from aiogram import types, Router, F
from aiogram.filters import CommandStart, and_f
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile
from asgiref.sync import sync_to_async
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from data.file_extensions import DocumentTypeFilter, is_valid_passport, yellow_fill, is_valid_pinfl, \
    contains_prohibited_product, red_fill, replace_words, get_all_passport, orange_fill
from data.file_extensions import is_phone_word_validator, blue_fill, violet_fill
from data.prohibit_product import REPLACE_WORDS
from keyboards import default_kb
from loader import dp, bot
from services.wb_parser.wb_parser import get_metadata_from_wb
from utils.utils import download_file, is_convertible_to_int

default_router = Router(name=__name__)


class ProhibitData:
    # получаем text и контакт для связи
    def __init__(self):
        self.product_name = None


prohibit_data = {}


# Создаем функцию для инициализации get_help_text
@sync_to_async
def get_prohibit_text(user_id):
    if user_id not in prohibit_data:
        prohibit_data[user_id] = ProhibitData()
    return prohibit_data[user_id]


@dp.message(
    lambda message: message.text == 'Отмена')
async def cmd_cancel(message: types.Message, state: FSMContext):
    """Отмена"""
    current_state = await state.get_state()
    if current_state:
        await state.clear()

    await message.answer(text='Выберите из списка', reply_markup=default_kb.menu)


@dp.message(CommandStart())
async def cmd_start(message: types.Message):
    await message.answer('Здравствуйте. Выберите из списка',
                         reply_markup=default_kb.menu)


@dp.message(F.text == 'Загрузить файл')
async def prompt_file_upload(message: types.Message):
    await message.answer('Вставьте файл', reply_markup=default_kb.cancel_markup)


@default_router.message(and_f(F.document, DocumentTypeFilter(allowed_extensions=['.xlsx', '.xls'])))
async def get_file_excel(message: types.Message, context: dict):
    # Скачивание файла
    file_stream, error = await download_file(bot, message.document)
    if error:
        await message.answer(error)
        return

    await message.answer('Файл скачен, идёт обработка. Подождите!!')

    # Загрузка данных из файла Excel
    df = pd.read_excel(file_stream, header=4)

    # Проверка на наличие нужных столбцов
    required_columns = ['ШК', 'Артикул сайта', 'Наименование товара', 'Описание', 'Баркод',
                        'ФИО получателя физ. лица',
                        'Номер паспорта', 'ТН ВЭД', 'Пинфл', 'Контактный номер', ]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        await message.answer(f"В файле отсутствуют столбцы: {', '.join(missing_columns)}")
        return

    # Преобразование ПИНФЛ в целые числа
    df['Пинфл'] = df['Пинфл'].astype('Int64')
    df['ТН ВЭД'] = df['ТН ВЭД'].astype('Int64')
    df.replace({np.nan: None}, inplace=True)

    # Проверяем, можно ли все элементы первой строки преобразовать в целые числа
    if all(map(is_convertible_to_int, df.iloc[0])):
        # Удаление первой строки, если она содержит только значения, преобразуемые в целые числа
        df.drop(index=0, inplace=True)
        df.reset_index(drop=True, inplace=True)

    # Получаем токен и Штрих коды с сохранением их порядка для загрузки данных из ВБ
    shk_dict = OrderedDict({str(int(shk)): None for shk in df['ШК'].tolist() if shk})
    token = context["access_token"]

    # Загрузка данных о товаре из WB и обновления токена авторизации
    parsed_df, new_token = await get_metadata_from_wb(message, shk_dict, token)
    context["access_token"] = new_token
    # Объединение данных из ВБ с данными из реестра
    df["Категория"] = parsed_df["Категория"]
    df["Подкатегория"] = parsed_df["Подкатегория"]
    df["Артикул сайта"] = parsed_df["Артикул"]
    # # для теста
    # df["Артикул"] = parsed_df["Артикул"]
    # df["Названия"] = parsed_df["Названия"]

    # Загрузка рабочего листа Excel
    file_stream = io.BytesIO()  # Сброс позиции для перезаписи файла
    df.to_excel(file_stream, index=False)
    file_stream.seek(0)  # Сброс позиции потока для openpyxl
    workbook = load_workbook(file_stream)
    sheet = workbook.active

    # Индексы колонок с паспортом и ПИНФЛ
    passport_col_idx = df.columns.get_loc('Номер паспорта') + 1
    pinfl_col_idx = df.columns.get_loc('Пинфл') + 1
    product_name_col_idx = df.columns.get_loc('Наименование товара') + 1
    description_col_idx = df.columns.get_loc('Описание') + 1

    duplicate_orders_dict = defaultdict(list)
    prohibited_product_rows = defaultdict(bool)

    unique_passports = set()
    unique_data = []
    # all_passport = get_all_passport()
    for index, row in df.iterrows():
        passport = str(row.get('Номер паспорта', '')).strip()
        pinfl = str(row.get('Пинфл', ''))
        hs_code = row.get('ТН ВЭД', 0)
        if hs_code is None or hs_code == '':
            hs_code = 0
        else:
            hs_code = int(hs_code)
        product_name = str(row.get('Наименование товара', ''))
        description = str(row.get('Описание', ''))
        barcode = row.get('Баркод', '')
        full_name = str(row.get('ФИО получателя физ. лица', ''))
        row_index = int(index + 2)

        product_name = replace_words(product_name, REPLACE_WORDS)
        description = replace_words(description, REPLACE_WORDS)
        df.at[index, 'Наименование товара'] = product_name
        df.at[index, 'Описание'] = description

        # Обновление значений в рабочем листе
        sheet.cell(row=row_index, column=product_name_col_idx, value=product_name)
        sheet.cell(row=row_index, column=description_col_idx, value=description)
        invalid_data = False

        if contains_prohibited_product(product_name) or contains_prohibited_product(description):
            # Помечаем строку красным цветом
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_index,
                                  column=col_idx)  # +7, так как header=4 и skiprows=[5] добавляют смещение
                cell.fill = red_fill

            # Сохраняем номер строки чтобы не перекрашивать её в  фиолетовый
            prohibited_product_rows[str(row_index)] = True
        is_valid, cleaned_passport = is_valid_passport(passport)
        if not is_valid:
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_index,
                                  column=col_idx)
                cell.fill = yellow_fill
            # cell = sheet.cell(row=row_index, column=passport_col_idx)
            # cell.fill = yellow_fill
            # invalid_data = True
        else:
            # Внесение изменений в DataFrame и лист Excel
            df.at[index, 'Номер паспорта'] = cleaned_passport
            sheet.cell(row=row_index, column=passport_col_idx, value=cleaned_passport)
        """ закрашиваем перелимиты"""
        # if cleaned_passport in all_passport:
        #     for col_idx in range(1, len(df.columns) + 1):
        #         cell = sheet.cell(row=row_index,
        #                           column=col_idx)
        #         cell.fill = orange_fill

        if not is_valid_pinfl(pinfl):
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_index,
                                  column=col_idx)
                cell.fill = yellow_fill

        max_col_idx = sheet.max_column
        if is_phone_word_validator(hs_code):
            for col_idx in range(1, max_col_idx + 1):
                cell = sheet.cell(row=row_index, column=col_idx)
                cell.fill = blue_fill
            invalid_data = True

        # Считаем кол-во заказов по пользователям, нужно чтобы потом красить повторяющие красить
        if passport and pinfl and barcode and not prohibited_product_rows.get(str(row_index), None):
            duplicate_orders_dict[f'{passport}_{pinfl}_{barcode}'].append(row_index)

        # Сохранение уникальных паспортов и данных, только если данные корректны
        if passport not in unique_passports and not invalid_data:
            unique_passports.add(passport)
            unique_data.append({
                'Номер паспорта': passport,
                'ФИО': full_name,
                'Пинфл': int(pinfl) if isinstance(pinfl, (int, float)) else pinfl
            })
    # Красив на фиолетовый повторяющиеся заказы одного пользователя
    for key in duplicate_orders_dict.keys():
        duplicate_rows = duplicate_orders_dict[key]
        if len(duplicate_rows) > 3:
            for row in duplicate_rows:
                for cell in sheet[row]:
                    cell.fill = violet_fill

    # Создание нового листа для уникальных данных
    unique_df = pd.DataFrame(unique_data)
    unique_sheet = workbook.create_sheet(title="Уникальные данные")

    for r_idx, row in enumerate(dataframe_to_rows(unique_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            unique_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Сохранение изменений в новый файл Excel в память
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    file = BufferedInputFile(output_stream.getvalue(), "output.xlsx")
    await message.answer_document(file)


@dp.message(F.text == 'Добавить запрещёнку')
async def added_prohibit_product(message: types.Message, state: FSMContext):
    await message.answer('В работе..', reply_markup=default_kb.cancel_markup)
    # await state.set_state(ProhibitProductDataState.product_name)


#
#
# @dp.message(ProhibitProductDataState.product_name)
# async def get_product_name(message: types.Message, state: FSMContext):
#     user_id = message.chat.id
#     prohibit_data = await get_prohibit_text(user_id)
#     prohibit_data.product_name = message.text
#     prohibit_data.user_id = message.chat.id
#     await message.answer(text=_(ru_texts['enter_contact_info'], selected_language),
#                          reply_markup=cancel_markup(user_id))
#     await state.set_state(HelpState.contact)
#
#
@dp.message(F.text == 'Список запрещённых товаров')
async def prohibit_product_list(message: types.Message):
    await message.answer('В работе..', reply_markup=default_kb.cancel_markup)
#
# async def save_prohibit_name(product_name, session_maker):
#     """
#     Сохраняем в бд
#     :param product_name:
#     :param session_maker:
#     :return:
#     """
#     country = await ProhibitedProduct.create_country(
#         product_name=product_name,
#         session_maker=session_maker)
#     return country
