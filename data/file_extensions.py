import os
import re

import gspread
from aiogram import types
from aiogram.filters import BaseFilter
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.styles import PatternFill

from data.config import PASSPORT_URL
from data.prohibit_product import PROHIBIT_PRODUCT, REPLACEMENT

# Задаем цвета для окрашивания
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
dark_gray_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
violet_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Путь к папке для загрузок
DOWNLOADS_DIR = 'downloads'
if not os.path.exists(DOWNLOADS_DIR):
    os.makedirs(DOWNLOADS_DIR)


class DocumentTypeFilter(BaseFilter):
    def __init__(self, allowed_extensions: list = None):
        self.allowed_extensions = ['.xlsx', '.xls'] if not allowed_extensions else allowed_extensions

    async def __call__(self, message: types.Message) -> bool:
        if message.document:
            file_extension = os.path.splitext(message.document.file_name)[1].lower()
            if file_extension in self.allowed_extensions:
                return True

        await message.answer("Неправильное расширение файла. Доступные форматы: "
                             f"<b>{', '.join(self.allowed_extensions)}</b>")
        return False


def get_all_passport(document_url: str = PASSPORT_URL):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('data/credentials.json', scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_url(document_url)
    worksheet = sheet.sheet1
    passports_column = worksheet.col_values(1)  # 1 означает первый столбец

    passports = [passport for passport in passports_column if passport]
    return passports


PASSPORT_CODE = ['AA', 'AB', 'AC', 'FA', 'KA', 'AD', 'FB', 'XS', 'FS', 'FB', 'FK']


# Функция для проверки корректности номера паспорта
def is_valid_passport(passport: str):
    # Удаляем пробелы и приводим буквы к заглавному регистру
    cleaned_passport = passport.strip().replace(' ', '').upper()
    if cleaned_passport[0:2] in PASSPORT_CODE:
        result = (isinstance(cleaned_passport, str) and
                  len(cleaned_passport) == 9 and
                  cleaned_passport[:2].isalpha() and
                  cleaned_passport[2:].isdigit())
    else:
        result = 0
    return result, cleaned_passport


def is_valid_pinfl(pinfl):
    return isinstance(pinfl, str) and len(pinfl) == 14 and pinfl.isdigit()


def is_phone_word_validator(value: int) -> bool:
    return 8517140000 == int(value)


def contains_prohibited_product(text):
    """Функция для проверки наличия запрещённого товара в тексте."""
    for product in PROHIBIT_PRODUCT:
        if product.lower() in text.lower():
            return True
    return False


def replace_words(text, replacements):
    pattern = re.compile('|'.join(re.escape(word) for word in replacements), re.IGNORECASE)
    return pattern.sub(lambda match: REPLACEMENT, text)


def highlight_invalid_cell(sheet, row_index, col_idx, fill_color):
    """Подсветить ячейку цветом."""
    max_col_idx = sheet.max_column
    for col_idx in range(1, max_col_idx +1):
        cell = sheet.cell(row=row_index, column=col_idx)
        cell.fill = fill_color