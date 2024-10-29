import io
from collections import defaultdict, OrderedDict

import numpy as np
import pandas as pd
from aiogram.types import BufferedInputFile
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from data.file_extensions import is_phone_word_validator, blue_fill, violet_fill, dark_gray_fill
from data.file_extensions import is_valid_passport, yellow_fill, is_valid_pinfl, \
    contains_prohibited_product, red_fill, replace_words, get_all_passport, highlight_invalid_cell
from data.prohibit_product import REPLACE_WORDS
from services.passport_checker import PassportChecker
from services.wb_parser.wb_parser import get_metadata_from_wb
from utils.utils import is_convertible_to_int


async def check_required_columns(required_columns: list, df):
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return f"В файле отсутствуют столбцы: {', '.join(missing_columns)}"
    else:
        return True


# async def check

async def checker_info_from_wb(df, context, message):
    required_columns = ['ШК', 'Артикул сайта', 'Наименование товара', 'Описание', 'Баркод',
                        'ФИО получателя физ. лица',
                        'Номер паспорта', 'ТН ВЭД', 'Пинфл', 'Контактный номер', ]
    required = await check_required_columns(required_columns=required_columns, df=df)
    if required is not True:
        await message.answer(required)
        return
    # Преобразование ПИНФЛ в целые числа
    df['Пинфл'] = df['Пинфл'].astype('Int64')
    df['ТН ВЭД'] = df['ТН ВЭД'].astype('Int64')
    df.replace({np.nan: None}, inplace=True)

    # Проверяем, можно ли все элементы первой строки преобразовать в целые числа
    if all(map(is_convertible_to_int, df.iloc[0])):
        # Удаление первой строки, если она содержит только значения, преобразуемые в целые числа
        df.drop(index=0, inplace=True)
        df.reset_index(drop=True, inplace=True)

    # Получаем токен и Штрих коды с сохранением их порядка для загрузки данных из ВБ
    shk_dict = OrderedDict({str(int(shk)): None for shk in df['ШК'].tolist() if shk})
    token = context["access_token"]

    # Загрузка данных о товаре из WB и обновления токена авторизации
    parsed_df, new_token = await get_metadata_from_wb(message, shk_dict, token)
    context["access_token"] = new_token
    # Объединение данных из ВБ с данными из реестра
    df["Категория"] = parsed_df["Категория"]
    df["Подкатегория"] = parsed_df["Подкатегория"]
    df["Артикул сайта"] = parsed_df["Артикул"]
    # # для теста
    # df["Артикул"] = parsed_df["Артикул"]
    df["Названия"] = parsed_df["Названия"]

    # Загрузка рабочего листа Excel
    file_stream = io.BytesIO()  # Сброс позиции для перезаписи файла
    df.to_excel(file_stream, index=False)
    file_stream.seek(0)  # Сброс позиции потока для openpyxl
    workbook = load_workbook(file_stream)
    sheet = workbook.active

    # Индексы колонок с паспортом и ПИНФЛ
    passport_col_idx = df.columns.get_loc('Номер паспорта') + 1
    pinfl_col_idx = df.columns.get_loc('Пинфл') + 1
    product_name_col_idx = df.columns.get_loc('Наименование товара') + 1
    description_col_idx = df.columns.get_loc('Описание') + 1

    duplicate_orders_dict = defaultdict(list)
    prohibited_product_rows = defaultdict(bool)

    unique_passports = set()
    unique_data = []
    for index, row in df.iterrows():
        passport = str(row.get('Номер паспорта', '')).strip()
        pinfl = str(row.get('Пинфл', ''))
        hs_code = row.get('ТН ВЭД', 0)
        if hs_code is None or hs_code == '':
            hs_code = 0
        else:
            hs_code = int(hs_code)
        product_name = str(row.get('Наименование товара', ''))
        description = str(row.get('Описание', ''))
        barcode = row.get('Баркод', '')
        full_name = str(row.get('ФИО получателя физ. лица', ''))
        row_index = int(index + 2)

        product_name = replace_words(product_name, REPLACE_WORDS)
        description = replace_words(description, REPLACE_WORDS)
        df.at[index, 'Наименование товара'] = product_name
        df.at[index, 'Описание'] = description

        # Обновление значений в рабочем листе
        sheet.cell(row=row_index, column=product_name_col_idx, value=product_name)
        sheet.cell(row=row_index, column=description_col_idx, value=description)
        invalid_data = False

        if contains_prohibited_product(product_name) or contains_prohibited_product(description):
            # Помечаем строку красным цветом
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_index,
                                  column=col_idx)  # +7, так как header=4 и skiprows=[5] добавляют смещение
                cell.fill = red_fill

            # Сохраняем номер строки чтобы не перекрашивать её в  фиолетовый
            prohibited_product_rows[str(row_index)] = True
        if not is_valid_pinfl(pinfl):
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
            invalid_data = True

        is_valid, cleaned_passport = is_valid_passport(passport=passport)
        if is_valid == 0:  # Паспорт не принадлежит нашей стране
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, dark_gray_fill)
        elif is_valid is False:
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
            invalid_data = True
        else:
            check_in_service = PassportChecker().passport_pinfl_is_correct(pinfl=pinfl,
                                                                           passport_serial_number=passport)
            if check_in_service == 1:
                # Внесение изменений в DataFrame и лист Excel
                df.at[index, 'Номер паспорта'] = cleaned_passport
                sheet.cell(row=row_index, column=passport_col_idx, value=cleaned_passport)
            elif check_in_service == 0:
                await message.answer(
                    'Сервис по определению ПД недоступен!!,\n обратитесь к Админу https://t.me/Razzakov_Timur')
            else:
                highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
                invalid_data = True

        max_col_idx = sheet.max_column
        if is_phone_word_validator(hs_code):
            for col_idx in range(1, max_col_idx + 1):
                cell = sheet.cell(row=row_index, column=col_idx)
                cell.fill = blue_fill
            invalid_data = True

        # Считаем кол-во заказов по пользователям, нужно чтобы потом красить повторяющие красить
        if passport and pinfl and barcode and not prohibited_product_rows.get(str(row_index), None):
            duplicate_orders_dict[f'{passport}_{pinfl}_{barcode}'].append(row_index)

        # Сохранение уникальных паспортов и данных, только если данные корректны
        if passport not in unique_passports and not invalid_data:
            unique_passports.add(passport)
            unique_data.append({
                'Номер паспорта': passport,
                'ФИО': full_name,
                'Пинфл': int(pinfl) if isinstance(pinfl, (int, float)) else pinfl
            })
    # Красив на фиолетовый повторяющиеся заказы одного пользователя
    for key in duplicate_orders_dict.keys():
        duplicate_rows = duplicate_orders_dict[key]
        if len(duplicate_rows) > 3:
            for row in duplicate_rows:
                for cell in sheet[row]:
                    cell.fill = violet_fill

    # Создание нового листа для уникальных данных
    unique_df = pd.DataFrame(unique_data)
    unique_sheet = workbook.create_sheet(title="Уникальные данные")

    for r_idx, row in enumerate(dataframe_to_rows(unique_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            unique_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Сохранение изменений в новый файл Excel в память
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    file = BufferedInputFile(output_stream.getvalue(), "output_wb.xlsx")
    return file


async def checker_info_from_ozone(df, context, message):
    required_columns = ['Штрих код', 'Артикул', 'Описание на русском', 'Описание на английском',
                        'Фамилия', 'Имя', 'Отчество', 'Количество',
                        'Номер паспорта', 'Код ТНВЭД', 'ИНН', 'Телефон', ]
    required = await check_required_columns(required_columns=required_columns, df=df)
    if required is not True:
        await message.answer(required)
        return

    file_stream = io.BytesIO()  # Сброс позиции для перезаписи файла
    df.to_excel(file_stream, index=False)
    file_stream.seek(0)  # Сброс позиции потока для openpyxl
    workbook = load_workbook(file_stream)
    sheet = workbook.active

    # Индексы колонок с паспортом и ПИНФЛ
    passport_col_idx = df.columns.get_loc('Номер паспорта') + 1
    pinfl_col_idx = df.columns.get_loc('ИНН') + 1
    description_col_idx = df.columns.get_loc('Описание на русском') + 1

    duplicate_orders_dict = defaultdict(list)
    prohibited_product_rows = defaultdict(bool)

    unique_passports = set()
    unique_data = []
    for index, row in df.iterrows():
        passport = str(row.get('Номер паспорта', '')).strip()
        pinfl = str(row.get('ИНН', ''))
        hs_code = row.get('Код ТНВЭД', 0)
        if hs_code is None or hs_code == '':
            hs_code = 0
        else:
            hs_code = int(hs_code)
        description = str(row.get('Описание на русском', ''))
        full_name = f"{str(row.get('Фамилия', '')), str(row.get('Имя', '')), str(row.get('Отчество', ''))}"
        row_index = int(index + 2)

        description = replace_words(description, REPLACE_WORDS)
        df.at[index, 'Описание на русском'] = description

        # Обновление значений в рабочем листе
        sheet.cell(row=row_index, column=description_col_idx, value=description)
        invalid_data = False

        if contains_prohibited_product(description):
            # Помечаем строку красным цветом
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_index,
                                  column=col_idx)
                cell.fill = red_fill
            # Сохраняем номер строки чтобы не перекрашивать её
            prohibited_product_rows[str(row_index)] = True
        if not is_valid_pinfl(pinfl):
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
            invalid_data = True

        is_valid, cleaned_passport = is_valid_passport(passport=passport)
        if is_valid == 0:  # Паспорт не принадлежит нашей стране
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, dark_gray_fill)
        elif is_valid is False:
            highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
            invalid_data = True
        else:
            check_in_service = PassportChecker().passport_pinfl_is_correct(pinfl=pinfl,
                                                                           passport_serial_number=passport)
            if check_in_service == 1:
                # Внесение изменений в DataFrame и лист Excel
                df.at[index, 'Номер паспорта'] = cleaned_passport
                sheet.cell(row=row_index, column=passport_col_idx, value=cleaned_passport)
            elif check_in_service == 0:
                await message.answer(
                    'Сервис по определению ПД недоступен!!,\n обратитесь к Админу https://t.me/Razzakov_Timur')
            else:
                highlight_invalid_cell(sheet, row_index, pinfl_col_idx, yellow_fill)
                invalid_data = True

        max_col_idx = sheet.max_column
        if is_phone_word_validator(hs_code):
            for col_idx in range(1, max_col_idx + 1):
                cell = sheet.cell(row=row_index, column=col_idx)
                cell.fill = blue_fill
            invalid_data = True

        # Считаем кол-во заказов по пользователям, нужно чтобы потом красить повторяющие красить
        if passport and pinfl and hs_code and not prohibited_product_rows.get(str(row_index), None):
            duplicate_orders_dict[f'{passport}_{pinfl}_{hs_code}'].append(row_index)

        # Сохранение уникальных паспортов и данных, только если данные корректны
        if passport not in unique_passports and not invalid_data:
            unique_passports.add(passport)
            unique_data.append({
                'Номер паспорта': passport,
                'ФИО': full_name,
                'Пинфл': int(pinfl) if isinstance(pinfl, (int, float)) else pinfl
            })
    # Красив на фиолетовый повторяющиеся заказы одного пользователя
    for key in duplicate_orders_dict.keys():
        duplicate_rows = duplicate_orders_dict[key]
        if len(duplicate_rows) > 3:
            for row in duplicate_rows:
                for cell in sheet[row]:
                    cell.fill = violet_fill

    # Создание нового листа для уникальных данных
    unique_df = pd.DataFrame(unique_data)
    unique_sheet = workbook.create_sheet(title="Уникальные данные")

    for r_idx, row in enumerate(dataframe_to_rows(unique_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            unique_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Сохранение изменений в новый файл Excel в память
    output_stream = io.BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    file = BufferedInputFile(output_stream.getvalue(), "output_ozone.xlsx")
    return file
